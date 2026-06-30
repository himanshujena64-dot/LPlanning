# ═══════════════════════════════════════════════════════════════
# EXCEL TEMPLATE GENERATOR
# Paste this whole block into your app, right after the HELPERS
# section (after the `parse_all_month_cols` / `standardize_req_header`
# functions, before `detect_req_header_row` is fine, or anywhere
# above the "SIDEBAR NAV" section).
# ═══════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

_TPL_HEADER_FILL = PatternFill("solid", fgColor="1A6EF7")
_TPL_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TPL_NOTE_FONT   = Font(color="9CA3AF", italic=True, size=10)
_TPL_SAMPLE_FONT = Font(color="6B7280", size=10)

def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _TPL_HEADER_FILL
        cell.font = _TPL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _add_note(ws, text, row, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 4))
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _TPL_NOTE_FONT

def build_template_bytes(kind: str) -> bytes:
    """
    Build a downloadable blank/sample XLSX template for the given file kind.
    kind: 'bom' | 'mb52' | 'prod_plan' | 'receipt' | 'segment' | 'aging'
    """
    wb = Workbook()
    ws = wb.active

    if kind == "bom":
        ws.title = "BOM"
        headers = ["BOM Header", "BOM header descripti", "Alt", "Level", "Parent",
                   "Component", "Component descriptio", "Required Qty",
                   "Base unit", "Procurement type", "Special procurement"]
        ws.append(headers)
        _style_header(ws, len(headers))
        sample = [
            ["FG0001", "Outdoor Unit 1.5T", "0", 1, "FG0001", "ASSY0001", "Compressor Assembly", 1, "EA", "E", ""],
            ["FG0001", "Outdoor Unit 1.5T", "0", 2, "ASSY0001", "PHN0001", "Phantom Sub-Assembly", 1, "EA", "L", "50"],
            ["FG0001", "Outdoor Unit 1.5T", "0", 3, "PHN0001", "0010748458", "Copper Tube 6mm", 2, "M", "F", ""],
            ["FG0001", "Outdoor Unit 1.5T", "0", 4, "0010748458", "0010300601DEL", "Bracket Mount", 4, "EA", "F", ""],
        ]
        for r in sample: ws.append(r)
        for r in range(2, 6):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = _TPL_SAMPLE_FONT
        _add_note(ws, "Note: 'Special procurement' = your Phantom code (default 50) marks pass-through assemblies. "
                       "Delete sample rows (2-5) before filling real data. 'Level' must start at 1 under each BOM Header.",
                  row=7, ncols=len(headers))
        _autofit(ws, [14, 26, 6, 7, 14, 16, 26, 12, 10, 16, 18])

    elif kind == "mb52":
        ws.title = "MB52 Stock"
        headers = ["Material", "Unrestricted", "Quality Inspection"]
        ws.append(headers)
        _style_header(ws, len(headers))
        sample = [
            ["0010748458", 1200, 50],
            ["0010748460", 340, 0],
            ["0010300601DEL", 980, 20],
        ]
        for r in sample: ws.append(r)
        for r in range(2, 5):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = _TPL_SAMPLE_FONT
        _add_note(ws, "Note: This is normally a SAP MB52 TXT export (tab-delimited). You may also save this sheet "
                       "as .xlsx and the engine will read 'Material' + sum of Unrestricted + Quality Inspection columns. "
                       "Delete sample rows before filling real data.", row=6, ncols=len(headers))
        _autofit(ws, [18, 14, 18])

    elif kind == "prod_plan":
        ws.title = "Production Plan"
        today = date(2026, 5, 1)
        day_cols = [today + timedelta(days=i) for i in range(0, 10)]
        headers = ["Type", "FG_Com"] + [d.strftime("%Y-%m-%d") for d in day_cols]
        ws.append(headers)
        _style_header(ws, len(headers))
        sample = [
            ["FERT", "FG0001"] + [10, 10, 0, 15, 15, 0, 0, 20, 20, 10],
            ["FERT", "FG0002"] + [5, 5, 5, 0, 0, 10, 10, 0, 5, 5],
        ]
        for r in sample: ws.append(r)
        for r in range(2, 4):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = _TPL_SAMPLE_FONT
        _add_note(ws, "Note: Add one column per production date (any number of days/months). 'FG_Com' = finished "
                       "goods code matching 'BOM Header' in the BOM file. Delete sample rows before filling real data.",
                  row=5, ncols=len(headers))
        _autofit(ws, [10, 14] + [13] * len(day_cols))

    elif kind == "receipt":
        ws.title = "Receipts"
        headers = ["Component", "May-26", "Jun-26", "Jul-26", "Aug-26"]
        ws.append(headers)
        _style_header(ws, len(headers))
        sample = [
            ["0010748458", 100, 50, 0, 200],
            ["0010748460", 0, 30, 30, 0],
        ]
        for r in sample: ws.append(r)
        for r in range(2, 4):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = _TPL_SAMPLE_FONT
        _add_note(ws, "Note: First column = component code. Remaining columns = one per month (e.g. 'Jun-26'), "
                       "containing expected/received quantity for that month. Add/remove month columns as needed.",
                  row=5, ncols=len(headers))
        _autofit(ws, [18, 10, 10, 10, 10])

    elif kind == "segment":
        ws.title = "Import Part List"
        headers = ["Import Part", "RM Group"]
        ws.append(headers)
        _style_header(ws, len(headers))
        sample = [
            ["0010748458", "Copper"],
            ["0010748460", "Compressor"],
            ["0010300601DEL", "Sheet Metal"],
        ]
        for r in sample: ws.append(r)
        for r in range(2, 5):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = _TPL_SAMPLE_FONT
        _add_note(ws, "Note: list every import-constrained part code. 'RM Group' is optional (used for filtering).",
                  row=6, ncols=len(headers))
        _autofit(ws, [18, 16])

        ws2 = wb.create_sheet("Segment")
        headers2 = ["Segment", "FG_Code", "IDU", "Compatible_ODU"]
        ws2.append(headers2)
        _style_header(ws2, len(headers2))
        sample2 = [
            ["Premium", "FG0001", "IDU0001", "FG0002"],
            ["Standard", "FG0003", "IDU0002", "FG0004"],
        ]
        for r in sample2: ws2.append(r)
        for r in range(2, 4):
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=r, column=c).font = _TPL_SAMPLE_FONT
        _add_note(ws2, "Note: 'IDU' and 'Compatible_ODU' codes must match 'BOM Header' codes in the BOM file.",
                  row=5, ncols=len(headers2))
        _autofit(ws2, [14, 14, 14, 16])

    elif kind == "aging":
        ws.title = "Aging Material Details"
        headers = ["Material", "Material Description", "Material Type", "MAP",
                   "0-15 Qty", "16-30 Qty", "31-60 Qty", "61-90 Qty", "91-120 Qty",
                   "121-150 Qty", "151-180 Qty", "181-360 Qty", "Over361 Qty",
                   "0-15 Value", "16-30 Value", "31-60 Value", "61-90 Value", "91-120 Value",
                   "121-150 Value", "151-180 Value", "181-360 Value", "Over361 Value"]
        ws.append(headers)
        _style_header(ws, len(headers))
        sample = [
            ["0010748458", "Copper Tube 6mm", "ROH", 120, 50, 30, 0, 0, 100, 0, 0, 0, 0,
             6000, 3600, 0, 0, 12000, 0, 0, 0, 0],
        ]
        for r in sample: ws.append(r)
        for c in range(1, len(headers) + 1):
            ws.cell(row=2, column=c).font = _TPL_SAMPLE_FONT
        _add_note(ws, "Note: this matches the SAP aging report layout. One row per Material per Storage Location is "
                       "fine — the app consolidates automatically. 'MAP' = Moving Average Price. Delete sample row "
                       "before filling real data.", row=4, ncols=len(headers))
        _autofit(ws, [14, 26, 12, 10] + [11] * 18)

    else:
        raise ValueError(f"Unknown template kind: {kind}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def template_download_button(label, kind, filename, key):
    """Small, unobtrusive 'Download template' link-style button."""
    st.download_button(
        f"⬇ {label}",
        data=build_template_bytes(kind),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=key,
    )
